import csv
import io
import logging
import os
import time

from django.conf import settings
from django.http import HttpResponse, Http404
from django.shortcuts import render
from openpyxl import Workbook

from parsing_methods.ds4Parsing import BS4Parsing
from parsing_methods.request_letyshops_parsing import RequestsLetyShopsParsing
from parsing_methods.requestsParsing import RequestsParsing
from parsing_methods.scrapyParsing.scrapyParsing.spiders.arcady_spider import MainClassForScrapy
from parsing_methods.webDriverParsing import WebDriverParsing
from .models import Shop, Timer

logger = logging.getLogger(__name__)

file_name_for_shops_excel = "E:/Документы/PyCharmProject/Django/files/shops.xlsx"
file_name_for_shops_csv = "E:/Документы/PyCharmProject/Django/files/shops.csv"
file_name_for_times_excel = "E:/Документы/PyCharmProject/Django/files/times.xlsx"


# Create your views here.
def index(request):
    return HttpResponse("Hello! If you're admin, please, log in")


def shops(request):
    shop_list = get_shops_from_database()
    context = {'shop_list': shop_list}
    return render(request, 'shops.html', context)


def get_shops_from_database():
    return Shop.objects.order_by('-name')


def get_times_from_database():
    return Timer.objects.order_by('-time')


def parse(request):
    global method
    Shop.objects.all().delete()
    Timer.objects.all().delete()
    methods = [
        RequestsParsing(),
        BS4Parsing(),
        WebDriverParsing(),
        RequestsLetyShopsParsing(),
    ]
    shops_list = []
    for method in methods:
        try:
            start_time = time.time()
            shops_list.append(method.parsing())
            timer = Timer(name=method.get_name_class(),
                          time=(time.time() - start_time))
            timer.save()
        except Exception as e:
            logger.error("Ошибка при проходе методов : " + e.__str__())
    # try:
    #     run_scrapy_and_save_data()
    # except Exception as e:
    #     logger.error("Ошибка при запуске Scrapy: " + e.__str__())
    context = {'shop_list': save_shops_in_database(shops_list)}
    return render(request, 'shops.html', context)


def run_scrapy_and_save_data():
    start_time = time.time()
    spider = MainClassForScrapy()
    timer = Timer(name=MainClassForScrapy.__name__.__str__(),
                  time=(time.time() - start_time))
    timer.save()
    save_shops_in_database_for_scrapy(spider)


def save_shops_in_database_for_scrapy(spider):
    shops_list = spider.get_data_from_json()
    for shop in shops_list:
        new_shop = Shop(name=shop["name"],
                        discount=shop["discount"],
                        label=shop["label"],
                        image=shop["image"],
                        url=shop["url"])
        new_shop.save()


def save_shops_in_database(shops):
    shop_list = []
    for shop in shops:
        for item in shop:
            shop_list.append(item)
    for shop in shop_list:
        try:
            new_shop = Shop(name=shop.name, discount=shop.discount, label=shop.label, url=shop.url, image=shop.image)
            new_shop.save()
        except AttributeError as e:
            continue


def download_shops_in_excel(request):
    save_shops_in_excel()
    file_path = os.path.join(settings.MEDIA_ROOT, file_name_for_shops_excel)
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response
    raise Http404


def download_times_in_excel(request):
    save_times_in_excel()
    file_path = os.path.join(settings.MEDIA_ROOT, file_name_for_times_excel)
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response
    raise Http404

def download_shops_in_csv(request):
    save_shops_in_csv()
    file_path = os.path.join(settings.MEDIA_ROOT, file_name_for_shops_csv)
    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response
    raise Http404


def save_shops_in_excel():
    workbook = Workbook()
    sheet = workbook.create_sheet('Sheet1', 0)
    worksheet = workbook['Sheet1']
    worksheet['A1'] = 'НАЗВАНИЕ'
    worksheet['B1'] = 'ЗНАЧЕНИЕ КЭШБЭКА'
    worksheet['C1'] = 'ВАЛЮТА'
    worksheet['D1'] = 'КАРТИНКА МАГАЗИНА'
    worksheet['E1'] = 'АДРЕС МАГАЗИНА'
    shop_list = get_shops_from_database()
    for i, shop in enumerate(shop_list):
        index_of_row = i + 2
        worksheet['A' + str(index_of_row)] = shop.name
        worksheet['B' + str(index_of_row)] = shop.discount
        worksheet['C' + str(index_of_row)] = shop.label
        worksheet['D' + str(index_of_row)] = shop.image
        worksheet['E' + str(index_of_row)] = shop.url
    workbook.save(file_name_for_shops_excel)
    workbook.close()


def save_times_in_excel():
    workbook = Workbook()
    sheet = workbook.create_sheet('Sheet1', 0)
    worksheet = workbook['Sheet1']
    worksheet['A1'] = 'ИМЯ МЕТОДА'
    worksheet['B1'] = 'ВРЕМЯ ВЫПОЛНЕНИЯ'
    worksheet['C1'] = 'ДАТА ЗАМЕРА'
    time_list = get_times_from_database()
    for i, timer in enumerate(time_list):
        index_of_row = i + 2
        worksheet['A' + str(index_of_row)] = timer.name
        worksheet['B' + str(index_of_row)] = timer.time
        worksheet['C' + str(index_of_row)] = timer.date
    workbook.save(file_name_for_times_excel)
    workbook.close()


def save_shops_in_csv():
    shop_list = get_shops_from_database()
    with io.open(file_name_for_shops_csv, "w", newline="", encoding='utf8') as file:
        columns = ['name', 'discount', 'label', 'image', 'url']
        writer = csv.DictWriter(file, fieldnames=columns)
        for shop in shop_list:
            new_shop = {'name': shop.name,
                        'discount': shop.discount,
                        'label': shop.label,
                        'image': shop.image,
                        'url': shop.url}
            writer.writerow(new_shop)
